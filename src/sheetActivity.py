from selenium import webdriver
from selenium.webdriver.support.select import Select
import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.alert import Alert
from selenium.common.exceptions import NoAlertPresentException
from selenium.common.exceptions import UnexpectedAlertPresentException
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import *
from selenium.common.exceptions import NoAlertPresentException
from datetime import date
from datetime import time
from datetime import datetime
from selenium.common.exceptions import TimeoutException
import xlrd
from json import dumps
from selenium.common.exceptions import StaleElementReferenceException
from easygui import *
import logging
from src.config import AppConfiguration # import error if not for Src.
from src.auth import clickonlogin, askforlogincred

appConfig = AppConfiguration()
config = appConfig.getConfig()

def entersamplety(driver, sampleTy):
    if sampleTy:
        if sampleTy in ["Nasopharyngeal & Oropharyngeal", "Nasopharyngeal swab", "Oropharyngeal swab", "Nasal swab",
                        "Throat swab", "Sputum", "BAL", "ETA"]:
        # click on the sample type dropdown
            driver.find_element_by_id("sample_type").click()
            driver.implicitly_wait(3)
            # create a select class for that dropdown and click on the value
            sampleType = Select(driver.find_element_by_id("sample_type"))
            sampleType.select_by_visible_text(sampleTy)
        else:
            print("Incorrect sample type provided type for SRF ID:", SRFid)
            logging.info("Incorrect sample type provided type for SRF ID: " + str(SRFid))
    else:
        print("no sample type provided type for SRF ID:", SRFid)
        logging.info("No sample type provided type for SRF ID: " + str(SRFid))


def opensourcefile():
    # opens the excel sheet and count the number of rows and columns
    workbook = openpyxl.load_workbook(config["datafilepath"])
    # sheet = workbook.active #(if you have one sheet in the excel)
    sheet = workbook["DataEntry_Paste as values"]
    rows = sheet.max_row
    # note: data entry is from the last row of the sheet
    currentrow = rows
    print("Total number of rows : ", rows)
    logging.info("Total number of rows : " + str(rows))
    logging.info("Total number of rows for data entry : " + str(rows - 1))
    print("Total number of rows for data entry : ", rows - 1)
    cols = sheet.max_column
    print("cols : ", cols)
    logging.info("Total number of cols : " + str(cols) + '\n')
    return currentrow, sheet


def enterlogincredentials():
    if config["driverTy"] == "firefox":
        #path = cwd + '\\src\\drivers\\firefoxdriver.exe'       # path format corrected
        path = config["geckodriverpath"]
        driver = webdriver.Firefox(executable_path=path)
    else:
        #path = cwd + '\\src\\drivers\\chromedriver.exe'         # path format corrected
        path = ["chromedriverpath"]
        driver = webdriver.Chrome(executable_path=path)
    driver.get(config["URL"])
    assert "ICMR" in driver.title
    driver.find_element_by_name("username").send_keys(config["u_name"])
    driver.find_element_by_name("passwd").send_keys(config["pwd"])
    logging.info("ICMR Page loaded at  " +
                 datetime.now().strftime('%d_%m_%Y__%H_%M_%S_%f'))
    return driver


def login():
    '''
    if boolbox(
            'Do you want to enter the User Name & Password? Please click Yes if you wish to change or enter the username and password',
            'Please click Yes if you wish to change or enter the username and password',
            ('[<F1>]Yes', '[<F2>]No'), None, '[<F2>]No', '[<F1>]Yes'):
        askforlogincred()
    '''
    logincheck = 0
    while logincheck == 0:
        driver = enterlogincredentials()
        if clickonlogin(driver):
            logincheck = 1
            # Wait for the login to happen. Manual login can be done, in which case disable the above 3 lines of code
            print("logged in at ", datetime.now())
            logging.info("logged in at " +
                         datetime.now().strftime('%d_%m_%Y__%H_%M_%S_%f'))
            # msgbox('login successful','login successful','OK')
            return driver
        else:
            if boolbox("Login Failed, please check login credentials." + '\n' + '\n' + ' Click Yes to Reenter credentials ' + '\n' + alert.text,
                       title='Login Failed, Click Yes to Reenter credentials', choices=('[Y]es', '[N]o'), image=None, default_choice='Yes',
                       cancel_choice='No'):
                alert = driver.switch_to.alert
                alert.accept()
                askforlogincred()
                continue
            else:
                alert.accept()
                driver.close()
                quit()
                exit()

def srfidmenu(driver, SRFid):
    try:
        driver.find_element_by_xpath("//section[@class ='sidebar']/ul/li[4]/a").click()
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.ID, "srf_id"))
        )
        driver.implicitly_wait(3)
        # click on search by srf and enter the srf value, click on search button
        driver.find_element_by_xpath("//*[@id='srf_id']").send_keys(SRFid)
    except StaleElementReferenceException:
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, "//section[@class ='sidebar']/ul/li[4]/a"))
        )
        srfidmenu(driver, SRFid)
    except UnexpectedAlertPresentException:
        if driver:
            driver.close()
        driver = login()
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, "//section[@class ='sidebar']/ul/li[4]/a"))
        )
        srfidmenu(driver, SRFid)
    except ElementNotInteractableException:
        #/ html / body / div[1] / header / nav / a
        try:
            driver.find_element_by_xpath("/html/body/div[1]/header/nav/a").click()
            srfidmenu(driver, SRFid)
        except StaleElementReferenceException:
            if driver:
                driver.close()
            driver = login()
            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "//section[@class ='sidebar']/ul/li[4]/a"))
            )
            srfidmenu(driver, SRFid)
        except UnexpectedAlertPresentException:
            if driver:
                driver.close()
            driver = login()
            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "//section[@class ='sidebar']/ul/li[4]/a"))
            )
            srfidmenu(driver, SRFid)

def opensrfid(sheet, currentrow, driver):

    textalert = ''
    SRFid=sheet.cell(row=currentrow, column=2).value
    #for testing only the commented code
    #SRFid = 29527000
    if SRFid is None:
        #currentrow=currentrow-1
        return "no SRF ID",textalert,SRFid
    else:
        SRFid = int(float(SRFid))
        logging.info('{' + str(SRFid) + "   SRF ID Start of ......")
        print('\n','Start of  :', SRFid)
        # Click on Add record from SRF Portal
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//section[@class ='sidebar']/ul/li[4]/a"))
        )
        srfidmenu(driver, SRFid)
        try:
            driver.find_element_by_id("btn").click()
            driver.implicitly_wait(1)
            WebDriverWait(driver, 6).until(EC.alert_is_present())
            alert = driver.switch_to.alert
            logging.info(str(SRFid) + " is already taken into ICMR Portal")
            logging.info(alert.text + '\n' + str(alert))
            print('\n',SRFid, " is already taken up into ICMR Portal")
            print(alert.text)
            print(alert)
            textalert = alert.text
            alert.accept()
            return "Alert!!", textalert, SRFid
        except TimeoutException:
            textalert = ''
            #wait for the webpage opens fetching the SRF ID
            try:
                WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.ID, "final_result_of_sample"))
                )
                logging.info(str(SRFid) + " is available and open ")
                print('\n', SRFid, " is available and open ")
                return False, textalert, SRFid
            except WebDriverException:
                alert = driver.switch_to.alert
                alert.accept()
                #WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "butto"))).click()
                return "Alert!!", textalert, SRFid
        except WebDriverException:
            WebDriverWait(driver, 3).until(EC.alert_is_present())
            alert = driver.switch_to.alert
            logging.info(str(SRFid) + " is already taken into ICMR Portal")
            logging.info(alert.text + '\n' + str(alert))
            print('\n', SRFid, " is already taken up into ICMR Portal")
            print(alert.text)
            print(alert)
            textalert = alert.text
            alert.accept()
            return "Alert!!", textalert, SRFid

def entertestkit(driver, testKit):
    #global testKit
    # Testing kit used
    # driver.execute_script(
    #   "document.getElementById('testing_kit_used').value+='" + testKit + "'")
    TestingKit = Select(driver.find_element_by_id("testing_kit_used"))
    TestingKit.select_by_visible_text(testKit)
    driver.implicitly_wait(7)


def enterfinalresult(driver,fiResult,SRFid):
    if driver.find_elements_by_id("final_result_of_sample")[-1].get_attribute("value") == fiResult:
        print("SRF ID", SRFid, " Final Result is", fiResult, '\n')
        logging.info(str(SRFid) + " Final Result is" + fiResult +'\n')
    else: 
        FinalResult = Select(driver.find_elements_by_name("final_result_of_sample")[-1])
        if fiResult in ["Positive", "Inconclusive", "Spillage", "Sample Rejected", "Antigen Positive"]:
            try:
                FinalResult.select_by_visible_text(fiResult)
                driver.implicitly_wait(3)
                WebDriverWait(driver, 5).until(EC.alert_is_present())
                alert_finalRes = driver.switch_to.alert
                print("SRF ID", SRFid, " has an alert  ", alert_finalRes.text,"\n")
                logging.info(str(SRFid) + "  Has an alert" + '\n' + alert_finalRes.text + '\n' + str(alert_finalRes))
                alert_finalRes.accept()
            except TimeoutException:
                enterfinalresult(driver,fiResult)
        elif fiResult in ["Negative", "Under Process", "TrueNAT Screening Positive", "TrueNAT Screening Negative", "Antigen Negative"]:
            FinalResult.select_by_visible_text(fiResult)
            enterfinalresult(driver,fiResult,SRFid)
        #driver.execute_script(
        #    "document.getElementById('final_result_of_sample').value+='" + fiResult + "'")
        #logging.info(" Final Result is:  " + fiResult + " for " + str(SRFid))
        # FinalResult = Select(driver.find_element_by_name("final_result_of_sample"))
        # FinalResult.select_by_visible_text(fiResult)
    
        
    
def enterResults(driver, Egen, SRFid, CtEgene, ORF1a,CtORF, RDRP_SGene, CtRDRP, fiResult):
    #global Egen, SRFid, CtEgene, ORF1a,CtORF, RDRP_SGene, CtRDRP, fiResult
    # E gene/N gene/TrueNAT
    # print("Egen :", Egen)
    if Egen:
        #driver.execute_script(
            #    "document.getElementById('final_result_of_sample').value+='" + fiResult + "'")
        #driver.execute_script(
         #  "document.getElementById('covid19_result_egen').value+='" + Egen + "'")
        Egene = Select(driver.find_element_by_id("covid19_result_egene"))
        Egene.select_by_visible_text(Egen)
        logging.info("E Gene is:  " + Egen)
        # CT value of E Gene
        if Egen == "Positive" and CtEgene:
            driver.find_element_by_id("ct_value_screening").send_keys(round(CtEgene))
            logging.info("E Gene Ct value is:  " + str(round(CtEgene)) + '\n')

    # ORF1a/ORF1b/N/N2 Gene
    if ORF1a:
        #driver.execute_script(
        #    "document.getElementById('orf1b_confirmatory').value+='" + ORF1a + "'")
        ORF = Select(driver.find_element_by_id("orf1b_confirmatory"))
        ORF.select_by_visible_text(ORF1a)
        logging.info('ORF1a Gene is:  ' + ORF1a)
        # Ct value of ORF
        if ORF1a == "Positive" and CtORF:
            driver.find_element_by_id("ct_value_orf1b").send_keys(round(CtORF))
            logging.info("ORF1b Gene Ct value is:  " + str(round(CtORF)) + '\n')
    # RdRp/S gene
    if RDRP_SGene:
        #driver.execute_script(
        # "document.getElementById('rdrp_confirmatory').value+='" + RDRP_SGene + "'")
        RdRp = Select(driver.find_element_by_id("rdrp_confirmatory"))
        RdRp.select_by_visible_text(RDRP_SGene)
        logging.info(" ORF1a Gene is:  " + RDRP_SGene)
        if RDRP_SGene == "Positive" and CtRDRP:
            driver.find_element_by_id("ct_value_rdrp").send_keys(round(CtRDRP))
            logging.info("RdRP Gene Ct value is:  " + str(round(CtRDRP)) + '\n')
    # final result of covid
    enterfinalresult(driver,fiResult,SRFid)

def enterPatientid(driver, patientId, SRFid):
    #global sampleId,SRFid
    # Patient ID Entry - check if the patient ID field is to be entered, if so, enter the value
    try:
        pid_element = driver.find_element_by_id("patient_id")
        if pid_element:
            driver.find_element_by_name("patient_id").send_keys(patientId)
            driver.implicitly_wait(3)
            # print("Patient ID",pid_element," already exists for",SRFid)
            # outputfile.write("Patient ID" + str(SRFid) + "      ")
        else:
            print("Patient ID", pid_element, " not found, already exists for", SRFid)
            logging.info("Patient ID already exists for this SRF")
            # driver.find_element_by_name("patient_id").send_keys(sampleId)
    except NoSuchElementException:
        # print(NoSuchElementException)
        print("Patient ID already exists for", SRFid)
        logging.info("Patient ID already exists for this SRF")

def hospitalizationEntry(driver, hospitalization):
    #global hospitalization, SRFid
    if driver.find_element_by_id("hospital").is_selected():
        hospitalizationInForm = driver.find_element_by_id("hospitalized").get_attribute("value")
        if not hospitalizationInForm:
            #hospitalize = datetime.strptime(onset_symptoms, '%Y/%m/%d %H:%M:%S').strftime('%d-%m-%Y %H:%M:%S')
            #driver.find_element_by_id('date_of_onset_of_symptoms').clear()
            if hospitalization:
                hospital = Select(driver.find_element_by_id("hospitalized"))
                hospital.select_by_visible_text(hospitalization)

def checksubmission(driver, SRFid):
    try:
        # driver.find_element_by_id("btn").click()
        driver.implicitly_wait(3)
        WebDriverWait(driver, 8).until(EC.presence_of_element_located((By.ID, "patient_id")))
        print("Submitted successfully")
        logging.info(str(SRFid) + "  Submission Successful" + '\n')
        return True
    except TimeoutException:
        try:
            WebDriverWait(driver, 8).until(EC.presence_of_element_located((By.ID, "contact_number")))
            print("Submitted successfully")
            logging.info(str(SRFid) + "  Submission Successful" + '\n')
            return True
        except TimeoutException:
            if boolbox('Please verify details, Check submission and click on continue',
                       'Please Verify details and click on submit to complete the data entry and then Click Continue',
                       ('[<F1>]Continue', '[<F2>]No'), None, '[<F1>]Yes', '[<F2>]No', ):
                checksubmission(driver, SRFid)
            else:
                logging.info(str(SRFid) + "  Submission might have failed, please check" + '\n')
                return False

def ptOccupation(driver,occupation):
    try:
        driver.find_element_by_id("patient_occupation")
        pt_occupation = Select(
            driver.find_element_by_id("patient_occupation"))
        pt_occupation.select_by_visible_text(occupation)
    except TimeoutException:
        test = 1
    except UnexpectedTagNameException:
        test = 2
        # Security Guards
        # Sanitation
        # Police
        # Health Care Worker

def modeOftransport(driver):
    modeOfTransport = Select(
        driver.find_element_by_id("mode_of_transport"))
    modeOfTransport.select_by_visible_text("Not Applicable")

def sampleCollectedfrom(driver,sampleCollectedFrom):
    if driver.find_element_by_id("community").is_selected():
        # community
        try:
            sample_colleted_from = driver.find_element_by_id(
                "sample_collected_from")
            sample_colleted_from_value = sample_colleted_from.get_attribute(
                "value")
            if sample_colleted_from != None:
                if sample_colleted_from_value != ['Containment Zone', 'Non-containment area', 'Point of entry']:
                    sample_collected__from = Select(
                        driver.find_element_by_id("sample_collected_from"))
                    driver.implicitly_wait(2)
                    # district.select_by_visible_text(Dist)
                    sample_collected__from.select_by_visible_text(
                        sampleCollectedFrom)
        except ElementClickInterceptedException:
            print("Sample collected from not clickable")
        except TimeoutException:
            print("No field found for Sample collected from")

def finalsubmit(driver, fiResult, SRFid):
    #global textalert
    #global fiResult, SRFid, alert, alert, textalert1, alert1
    textalert1 = None

    if fiResult in ["Positive", "Inconclusive", "Spillage", "Sample Rejected", "Antigen Positive", "Negative", "Under Process", "TrueNAT Screening Positive", "TrueNAT Screening Negative", "Antigen Negative"]:
    
    #if fiResult == "Jimasfds":
        textalert1 = None
        try:
            driver.find_element_by_id("btn").click()
            driver.implicitly_wait(3)
            WebDriverWait(driver, 5).until(EC.alert_is_present())
            alert1 = driver.switch_to.alert
            print("SRF ID", SRFid, " has an alert")
            logging.info(str(SRFid) + "  Has an alert" + '\n' + alert1.text + '\n' + str(alert1))
            logging.info(str(SRFid) + "  Submission might have failed, please check" + '\n')
            textalert1 = alert1.text
            if textalert1:
                boolboxoutput = boolbox(str(SRFid) +  "Has an alert" + '\n' + textalert1 + '\n' + '\n' + 'Please Check submission, Click OK on the alert and click on continue to continue',
                       'Please Verify details and click on the alert to complete the data entry and then Click Continue', ('[<F1>]Continue', '[<F2>]No'), None, '[<F1>]Continue', '[<F2>]No')
                if boolboxoutput:
                    logging.info(str(SRFid) + "  Manual correction done" + '\n')
                    return True,textalert1
                else:
                    return False, textalert1
            
            # outputfile.write('\n' + alert.text + "Patient ID already exists for this SRF")
            #print(alert)
            #alert.accept()
            #return False
        except TimeoutException:
            textalert1 = None
            submit = checksubmission(driver, SRFid)
            if submit:
                #outputfile.write("     " + str(SRFid) + "  Manual correction done" + '\n')
                return True, textalert1
            else:
               return False, textalert1
    else:
        textalert1 = None
        #driver.find_element_by_id("btn").click()
        driver.implicitly_wait(6)
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "sample_cdate")))
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "sample_cdate")))
        submit = checksubmission(driver, SRFid)
        if submit:
            logging.info(str(SRFid) + "  Submitted Successfully" + '\n')
            return True, textalert1
        else:
            
            return False, textalert1

def fetchrowvalues(sheet, currentrow):
    # Read all requried values from the excel sheet row
    #global sampleTy, sampleId, sampleRDate, sampleTDate, testKit, Egen, ORF1a, RDRP_SGene, fiResult, CtEgene, CtORF, CtRDRP
    #global labId, phone, ICMRHqID, State, Dist , patientname, patientId, hospitalization, sampleCDate, alert
    #global nationality, age, age_in, Gender, ArogyaSetuDownload, address, pincode, category, occupation, Mobilebelongsto
    #global symptomstatus, sampleCollectedFrom, fathersName

    sampleTy = sheet.cell(row=currentrow, column=3).value
    sampleId = sheet.cell(row=currentrow, column=4).value
    sampleRDate = sheet.cell(row=currentrow, column=5).value
    sampleTDate = sheet.cell(row=currentrow, column=6).value

    testKit = sheet.cell(row=currentrow, column=7).value
    Egen = ORF1a = RDRP_SGene = fiResult = None
    #testing kits yet to be corrected for checking against permitted values
    if sheet.cell(row=currentrow, column=8).value in ["Positive", "Negative", "Under Process", "Inconclusive/Spillage/Rejected"]:
        Egen = sheet.cell(row=currentrow, column=8).value
    if sheet.cell(row=currentrow, column=9).value in ["Positive", "Negative", "Under Process", "Inconclusive/Spillage/Rejected"]:
        ORF1a = sheet.cell(row=currentrow, column=9).value
    if sheet.cell(row=currentrow, column=10).value in ["Positive", "Negative", "Under Process",
                                                      "Inconclusive/Spillage/Rejected"]:
        RDRP_SGene = sheet.cell(row=currentrow, column=10).value
    if   sheet.cell(row=currentrow, column=11).value in ["Positive", "Negative", "Under Process", "Inconclusive", "Spillage", "Sample Rejected", "TrueNAT Screening Positive", "TrueNAT Screening Negative", "Antigen Positive", "Antigen Negative"]:
        fiResult = sheet.cell(row=currentrow, column=11).value
    #else:
        #fiResult = None

    CtEgene = sheet.cell(row=currentrow, column=12).value
    CtORF = sheet.cell(row=currentrow, column=13).value
    CtRDRP = sheet.cell(row=currentrow, column=14).value #N
    #CtRDRP = sheet.cell(row=currentrow, column=14).value
    labId = sheet.cell(row=currentrow, column=15).value
    if config["prefix"] and labId and (sampleId is None):
                #print("Sample ID is ", sampleId)
                sampleId = config["prefix"] + str(labId)
                #print("Sample ID now is ", sampleId)
                print("Sample ID is ",sampleId)
                logging.info('Sample ID is: ' + sampleId + '\n')

    if sheet.cell(row=currentrow, column=16).value:
        phone = sheet.cell(row=currentrow, column=16).value
    else:
        phone = None
    ICMRHqID = sheet.cell(row=currentrow, column=17).value
    State = sheet.cell(row=currentrow, column=18).value
    Dist = sheet.cell(row=currentrow, column=19).value
    patientname = sheet.cell(row=currentrow, column=20).value
    if sheet.cell(row=currentrow, column=21).value:
        patientId = sheet.cell(row=currentrow, column=21).value
    else:
        patientId = sampleId

    if sheet.cell(row=currentrow, column=22).value:
        if sheet.cell(row=currentrow, column=22).value in ["Yes","No"]:
            hospitalization = sheet.cell(row=currentrow, column=22).value
        else:
            hospitalization = 'No'
    else:
        hospitalization  = 'No'

    if sheet.cell(row=currentrow, column=23).value:
        sampleCDate = sheet.cell(row=currentrow, column=23).value
    else:
        sampleCDate = None


    if sheet.cell(row=currentrow, column=24).value:
        nationality = sheet.cell(row=currentrow, column=24).value
    else:
        nationality = None

    if sheet.cell(row=currentrow, column=25).value:
        age = sheet.cell(row=currentrow, column=25).value
    else:
        age = None

    if sheet.cell(row=currentrow, column=26).value:
        if sheet.cell(row=currentrow, column=26).value in ["Years", "Months", "Days"]:
            age_in = sheet.cell(row=currentrow, column=26).value
        else:
            age_in = None
    else:
        age_in = None

    Gender = None
    if sheet.cell(row=currentrow, column=27).value:
        if sheet.cell(row=currentrow, column=27).value in ["Male","Female", "Transgender", "M", "F","T"]:
            Gender = sheet.cell(row=currentrow, column=27).value


    if sheet.cell(row=currentrow, column=28).value:
        if sheet.cell(row=currentrow, column=28).value in ["Yes", "No"]:
            ArogyaSetuDownload = sheet.cell(row=currentrow, column=28).value
        else:
            ArogyaSetuDownload = None
    else:
        ArogyaSetuDownload = None

    if sheet.cell(row=currentrow, column=29).value:
        address = sheet.cell(row=currentrow, column=29).value
    else:
        address = None

    if sheet.cell(row=currentrow, column=30).value:
        pincode = sheet.cell(row=currentrow, column=30).value
    else:
        pincode = None

    Mobilebelongsto = None
    if sheet.cell(row=currentrow, column=31).value:
        if sheet.cell(row=currentrow, column=31).value in ["Patient", "Relative"]:
            Mobilebelongsto = sheet.cell(row=currentrow, column=31).value

    occupation = "Others"
    if sheet.cell(row=currentrow, column=32).value:
        if sheet.cell(row=currentrow, column=32).value in ["Health Care Worker","Police", "Sanitation", "Security Guards","Others" ]:
            occupation = sheet.cell(row=currentrow, column=32).value

    symptomstatus = "Asymptomatic"
    if sheet.cell(row=currentrow, column=33).value:
        if sheet.cell(row=currentrow, column=33).value in ["Symptomatic","Asymptomatic"]:
            symptomstatus = sheet.cell(row=currentrow, column=33).value

    sampleCollectedFrom =  "Point of entry"
    if sheet.cell(row=currentrow, column=34).value:
        if sheet.cell(row=currentrow, column=34).value in ["Non-containment area", "Containment Zone", "Point of entry"]:
            sampleCollectedFrom = sheet.cell(row=currentrow, column=34).value

    if sheet.cell(row=currentrow, column=35).value:
        fathersName = sheet.cell(row=currentrow, column=35).value
    else:
        fathersName = None

    if sheet.cell(row=currentrow, column=36).value:
        category = sheet.cell(row=currentrow, column=36).value
    else:
        category = None

    #alert = None

    return sampleTy, sampleId, sampleRDate, sampleTDate, testKit, Egen, ORF1a, RDRP_SGene, fiResult, CtEgene, CtORF, CtRDRP, \
           labId, phone, ICMRHqID, State, Dist , patientname, patientId, hospitalization, sampleCDate, \
           nationality, age, age_in, Gender, ArogyaSetuDownload, address, pincode, category, occupation, Mobilebelongsto,\
           symptomstatus, sampleCollectedFrom, fathersName

def vaccinedose1(driver,Vaccine_dose_1_Date):
    try:
        if driver.find_element_by_id("vaccine_dose_1").get_attribute("value") == "":
            if Vaccine_dose_1_Date:
                if isinstance(Vaccine_dose_1_Date, str):
                    Vaccine_dose_1_Date = datetime.strptime(Vaccine_dose_1_Date, '%Y/%m/%d')
                driver.execute_script(
                    "document.getElementById('vaccine_dose_1').value+='" + Vaccine_dose_1_Date.strftime(
                        '%d-%m-%Y') + "'")
    except NoSuchElementException:
        print(" Check for Vaccine first dose date field")
    except TimeoutException:
        print("Check for Vaccine first dose date field")

def vaccinedose2(driver,Vaccine_dose_2_Date):
    try:
        if driver.find_element_by_id("vaccine_dose_2").get_attribute("value") == "":
            if Vaccine_dose_2_Date:
                if isinstance(Vaccine_dose_2_Date, str):
                    Vaccine_dose_2_Date = datetime.strptime(Vaccine_dose_2_Date, '%Y/%m/%d')
                driver.execute_script(
                    "document.getElementById('vaccine_dose_2').value+='" + Vaccine_dose_2_Date.strftime(
                        '%d-%m-%Y') + "'")
    except NoSuchElementException:
        print(" Check for Vaccine second dose date field")
    except TimeoutException:
        print("Check for Vaccine second dose date field")

def vaccinetype(driver,VaccineType):
    try:
        if driver.find_element_by_id("vaccine_type").get_attribute("value") == "":
            if VaccineType in ["Covishield", "Covaxin"]:
                VaccineTy = Select(driver.find_element_by_id("vaccine_type"))
                if VaccineType == "Covishield":
                    VaccineType.select_by_visible_text("Covishield")
                elif VaccineType == "Covaxin":
                    VaccineType.select_by_visible_text("Covaxin")
    except NoSuchElementException:
        print(" Check for Vaccine type field")
    except TimeoutException:
        print("Check for Vaccine type field")

def vaccineentry(driver, VaccineRec, VaccineType, Vaccine_dose_1_Date, Vaccine_dose_2_Date):
    #VaccineReceivedinForm = driver.find_element_by_id("vaccine_recevied").get_attribute("value")
    try:
        if driver.find_element_by_id("vaccine_recevied").get_attribute("value") in [""]:
            if VaccineRec:
                VaccineSelect = Select(driver.find_element_by_id("vaccine_recevied"))
                if VaccineRec in ["y", "Y", "Yes", "yes", "YES"]:
                    VaccineSelect.select_by_visible_text("Yes")
                    vaccinetype(driver, VaccineType)
                    vaccinedose1(driver,Vaccine_dose_1_Date)
                    vaccinedose2(driver, Vaccine_dose_2_Date)
                if VaccineRec in ["n", "N", "NO", "no", "No"]:
                    VaccineSelect.select_by_visible_text("No")
        elif driver.find_element_by_id("vaccine_recevied").get_attribute("value") in ["y", "Y", "Yes", "yes", "YES"]:
            vaccinetype(driver, VaccineType)
            vaccinedose1(driver, Vaccine_dose_1_Date)
            vaccinedose2(driver, Vaccine_dose_2_Date)
    except NoSuchElementException:
        print(" Check for Vaccine received field")
    except TimeoutException:
        print("Check for Vaccine received field")


def enterdates(driver, sampleRDate, sampleTDate, sampleCDate):
    #global sampleRDate, sampleTDate, sampleCDate
    if isinstance(sampleTDate, str):
        sampleTDate = datetime.strptime(sampleTDate, '%Y/%m/%d %H:%M')
    if isinstance(sampleRDate, str):
        sampleRDate = datetime.strptime(sampleRDate, '%Y/%m/%d %H:%M')
    driver.execute_script(
        "document.getElementById('sample_rdate').value+='" + sampleRDate.strftime('%d-%m-%Y %H:%M:%S') + "'")
    # driver.find_element_by_id("sample_tdate").clear()
    driver.execute_script(
        "document.getElementById('sample_tdate').value+='" + sampleTDate.strftime('%d-%m-%Y %H:%M:%S') + "'")
    onset_symptoms = driver.find_element_by_id("date_of_onset_of_symptoms").get_attribute("value")
    if onset_symptoms:
        format_onset_symptoms = datetime.strptime(onset_symptoms, '%Y/%m/%d %H:%M:%S').strftime('%d-%m-%Y %H:%M:%S')
        driver.find_element_by_id('date_of_onset_of_symptoms').clear()
        driver.execute_script(
            "document.getElementById('date_of_onset_of_symptoms').value+='" + format_onset_symptoms + "'")

    if driver.find_element_by_id("hospital").is_selected():
        hospitalization_date = driver.find_element_by_id("hospitalization_date").get_attribute("value")
        if hospitalization_date:
            format_hospitalization_date = datetime.strptime(hospitalization_date, '%Y/%m/%d %H:%M:%S').strftime(
                '%d-%m-%Y %H:%M:%S')
            driver.find_element_by_id('hospitalization_date').clear()
            driver.execute_script(
                "document.getElementById('hospitalization_date').value+='" + format_hospitalization_date + "'")

    try:
        SampleCollD = driver.find_element_by_id("sample_cdate")
        sample_colldate = driver.find_element_by_id("sample_cdate").get_attribute("value")
        if SampleCollD and sample_colldate is not None:
            if sampleCDate is not None:
                if isinstance(sampleCDate, str):
                    sampleCDate = datetime.strptime(sampleCDate, '%Y/%m/%d')
                #driver.execute_script(
                 #   "document.getElementById('sample_cdate').value+='" + sampleCDate.strftime('%d-%m-%Y %H:%M:%S') + "'")
                driver.execute_script(
                   "document.getElementById('sample_cdate').value+='" + sampleCDate.strftime('%d-%m-%Y') + "'")
    except NoSuchElementException:
        # print(NoSuchElementException)
        print("Patient Collection Date already entered", SRFid)
        logging.info("Patient Collection date already entered in the form")

def searchMenuclick(driver):
    #global phone, ICMRHqId
    # click on search  patient
    try:
        driver.find_element_by_xpath("//section[@class ='sidebar']/ul/li[6]/a").click()
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "contact_number"))
        )
        #driver.implicitly_wait(5)

    except StaleElementReferenceException:
        WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "//section[@class ='sidebar']/ul/li[6]/a"))
            )
        searchMenuclick(driver)
    except ElementNotInteractableException:
        #/ html / body / div[1] / header / nav / a
        driver.find_element_by_xpath("/html/body/div[1]/header/nav/a").click()
        searchMenuclick(driver)

def stateDist(driver, State, Dist, statecheck):
    #global
    if statecheck != 0:
        try:
            driver.refresh()
            driver.implicitly_wait(10)
            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.ID, "state"))
            )
        except UnexpectedAlertPresentException:
            try:
                driver.switch_to.alert.accept()
            except NoAlertPresentException:
                justpass = 1

    try:
        '''
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.ID, "state"))
        )
        '''
        if driver.find_element_by_id("state").get_attribute("value") =='':
            statesel = Select(driver.find_element_by_id("state"))
            driver.implicitly_wait(3)
            statesel.select_by_visible_text(State)
            driver.implicitly_wait(10)

        if driver.find_element_by_id("district").get_attribute("value") =='':
            district = Select(driver.find_element_by_id("district"))
            #driver.implicitly_wait(2)
            try:
                district.select_by_visible_text(Dist)
                #driver.implicitly_wait(3)
            except NoSuchElementException:
                statecheck = statecheck+1
                stateDist(driver, State, Dist, statecheck)
    except NoSuchElementException:
        print( " This is a follow up case")
    except TimeoutException:
        print("This is a followup case")

def openBysearch(driver, State, Dist, ICMRHqID, patientId, phone, patientname, Gender, age, age_in):
    #global SRFid, sheet, workbook, phone, ICMRHqID, State, Dist, patientId, patientname, Gender, age, age_in
    
    searchMenuclick(driver)

    if State is None:
        State = 'KARNATAKA'

    if Dist is None:
        Dist = 'BELAGAVI'

    #stateDist(driver, State, Dist)
    stateDist(driver, State, Dist, 0)
    if ICMRHqID:
        if ICMRHqID != '-':
            ICMRHqID = int(ICMRHqID)
            driver.find_element_by_id("icmr_id").send_keys(ICMRHqID)

    if patientId is not None:
        driver.find_element_by_id("patient_id").send_keys(patientId)
    # driver.find_element_by_xpath("//*[@id='srf_id']").send_keys(SRFid)

    if phone:
        phone = int(phone)
        driver.find_element_by_id("contact_number").send_keys(phone)

    if patientname:
        driver.find_element_by_id("patient_name").send_keys(patientname)

    if Gender:
        genderType = Select(driver.find_element_by_id("gender"))
        if Gender in ["m", "M", "Male", "male", "MALE"]:
            genderType.select_by_visible_text("Male")
        if Gender in ["f", "F", "Female", "female", "FEMALE"]:
            genderType.select_by_visible_text("Female")
        if Gender == "Transgender":
            genderType.select_by_visible_text("Transgender")

    if age:
        driver.find_element_by_id("age").send_keys(age)

    if age_in:
        if age_in == "Years":
            driver.find_element_by_id("age_year").click()
        if age_in == "Months":
            driver.find_element_by_id("age_month").click()
        if age_in == "Days":
            driver.find_element_by_id("age_day").click()

    try:
        driver.find_element_by_id("btn").click()
        driver.implicitly_wait(4)
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "sample_id")))
    except TimeoutException:
        if not boolbox('Have you selected the appropriate Patient ID? Click No to skip to next entry','Please open the appropriate Patient ID for data entry. Click No to skip to next entry', ('[<F1>]Yes', '[<F2>]No'), None, '[<F1>]Yes', '[<F2>]No',):
            return False
        else:
            return True

def entersampleid(driver,sampleId):
    #global sampleId
    # if sampleTy has a value, check and if not none, enter the value
    driver.execute_script("document.getElementById('sample_id').value+='" + sampleId + "'")
    # driver.find_element_by_name("sample_id").send_keys(sampleId)
    # driver.implicitly_wait(5)


def checkNenterage_in(driver,age_in):
    try:
        if not driver.find_element_by_id("age_year").is_selected():
            if not driver.find_element_by_id("age_month").is_selected():
                if not driver.find_element_by_id("age_day").is_selected():
                    if age_in:
                        if age_in == "Years":
                            driver.find_element_by_id("age_year").click()
                            return "age_in updated to Years"
                        if age_in == "Months":
                            driver.find_element_by_id("age_month").click()
                            return "age_in updated to Months"
                        if age_in == "Days":
                            driver.find_element_by_id("age_day").click()
                            return "age_in updated to Days"
                        else:
                            return "age_in has incorrect value"
                    else:
                        return "age_in has no value"
                else:
                   return "SRF has already age_in selected for day"
            else:
                return "SRF has already age_in selected for month"
        else: return "SRF has already age_in selected for year"
    except NoSuchElementException:
        print( " This is a follow up case")
    except TimeoutException:
        print("This is a followup case")

def enterAge(driver,age):
    try:
        if not driver.find_element_by_id("age").get_attribute("value"):
            driver.find_element_by_id("age").send_keys(age)
    except NoSuchElementException:
        print( " This is a follow up case")
    except TimeoutException:
        print("This is a followup case")

def enterGender(driver,Gender):
    try:
        if driver.find_element_by_id("gender").get_attribute("value")=="":
            if Gender:
                genderType = Select(driver.find_element_by_id("gender"))
                if Gender in ["m", "M", "Male", "male", "MALE"]:
                    genderType.select_by_visible_text("Male")
                if Gender in ["f", "F", "Female", "female", "FEMALE"]:
                    genderType.select_by_visible_text("Female")
                if Gender in ["Transgender", "T", "transgender", "TRANSGENDER"] :
                    genderType.select_by_visible_text("Transgender")
    except NoSuchElementException:
        print( " This is a follow up case")
    except TimeoutException:
        print("This is a followup case")

def enterPhonenum(driver,phone):
    try:
        if driver.find_element_by_id("contact_number").get_attribute("value") == "":
            if phone:
                phone = int(phone)
                driver.find_element_by_id("contact_number").send_keys(phone)
    except NoSuchElementException:
        print( " This is a follow up case")
    except TimeoutException:
        print("This is a followup case")

def aarogyasetuApp(driver, ArogyaSetuDownload):
    try:
        if driver.find_element_by_id("aarogya_setu_app_downloaded").get_attribute("value") == "":
            if ArogyaSetuDownload:
                ArogyaSelect = Select(driver.find_element_by_id("aarogya_setu_app_downloaded"))
                if ArogyaSetuDownload in ["y", "Y", "Yes", "yes", "YES"]:
                    ArogyaSelect.select_by_visible_text("Yes")
                if ArogyaSetuDownload in ["n", "N", "NO", "no", "No"]:
                    ArogyaSelect.select_by_visible_text("No")
    except NoSuchElementException:
        print( " This is a follow up case")
    except TimeoutException:
        print("This is a followup case")

def enterAddress(driver,address):
    try:
        if address:
            if not driver.find_element_by_id("address").get_attribute("value"):
                driver.find_element_by_id("address").send_keys(address)
    except NoSuchElementException:
        print( " This is a follow up case")
    except TimeoutException:
        print("This is a followup case")

def enterPincode(driver,pincode):
    try:
        if pincode:
            if not driver.find_element_by_id("pincode").get_attribute("value"):
                driver.find_element_by_id("pincode").send_keys(pincode)
    except NoSuchElementException:
        print( " This is a follow up case")
    except TimeoutException:
        print("This is a followup case")

def enterNationality(driver, nationality):
    try:
        if nationality:
            if not driver.find_element_by_id("nationality").get_attribute("value"):
                driver.find_element_by_id("nationality").send_keys(nationality)
    except NoSuchElementException:
        print( " This is a follow up case")
    except TimeoutException:
        print("This is a followup case")

def patientCategoryselect(driver):
    try:
        if driver.find_element_by_id("community").is_selected():
            if not driver.find_element_by_id("ncat17").is_selected():
                if not driver.find_element_by_id("ncat1").is_selected():
                    if not driver.find_element_by_id("ncat3").is_selected():
                        if not driver.find_element_by_id("ncat4").is_selected():
                            driver.find_element_by_id("ncat17").click()

        elif driver.find_element_by_id("hospital").is_selected():
            if not driver.find_element_by_id("ncat18").is_selected():
                if not driver.find_element_by_id("ncat9").is_selected():
                    if not driver.find_element_by_id("ncat10").is_selected():
                        if not driver.find_element_by_id("ncat11").is_selected():
                            if not driver.find_element_by_id("ncat12").is_selected():
                                if not driver.find_element_by_id("ncat13").is_selected():
                                    if not driver.find_element_by_id("ncat14").is_selected():
                                        if not driver.find_element_by_id("ncat15").is_selected():
                                            driver.find_element_by_id("ncat18").click()
        else:
            driver.find_element_by_id("hospital").click()
            driver.implicitly_wait(3)
            WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "ncat18")))
            driver.find_element_by_id("ncat18").click()
    except NoSuchElementException:
        print( " This is a follow up case")
    except TimeoutException:
        print("This is a followup case")

def symptomStatus(driver,symptomstatus):
    if not driver.find_element_by_id("status").get_attribute("value"):
        if symptomstatus:
            symptom = Select(driver.find_element_by_id("status"))
            if symptomstatus in ["Asymptomatic", "asymptomatic"]:
                symptom.select_by_visible_text("Asymptomatic")
            if symptomstatus in ["Symptomatic", "symptomatic"]:
                symptom.select_by_visible_text("Symptomatic")

def enterMobilebelongsto(driver,Mobilebelongsto):
    try:
        if not driver.find_element_by_id("contact_number_belongs_to").get_attribute("value"):
            if Mobilebelongsto:
                mobilenobel = select(driver.find_element_by_id("contact_number_belongs_to"))
                if Mobilebelongsto == "Patient":
                    mobilenobel.select_by_visible_text("Patient")
                elif Mobilebelongsto == "Relative":
                    mobilenobel.select_by_visible_text("Relative")
    except NoSuchElementException:
        print( " This is a follow up case")
    except TimeoutException:
        print("This is a followup case")

def enterpatientName(driver, patientname):
    try:
        if not driver.find_element_by_id("patient_name").get_attribute("value"):
            if patientname:
                if len(patientname) >=4:
                    driver.find_element_by_id("nationality").send_keys(patientname)
    except NoSuchElementException:
        print( " This is a follow up case")
    except TimeoutException:
        print("This is a followup case")