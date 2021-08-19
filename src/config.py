import openpyxl
import datetime
from datetime import datetime
from easygui import *
import logging
import os
import os.path
from os import path
#import its
import platform
from src.auth import askforlogincred #ModuleNotFoundError: No module named 'auth' fixed

cwd = os.getcwd()


class AppConfiguration:
    appConfig = {
        "startTime": "",
        #"configPath": config_Path,
        "default_data_folder": "",
        "datafilepath": "",
        "loginfilepath": "",
        "driverTy": "",
        "logins": "",
        "URL": "https://cvstatus.icmr.gov.in",
        "u_name": "",
        "pwd": "",
        "loginsheet": "",
        "prefix": "",
        "ProceedWithRATfollowup": "",
        "labUniqueAlert": "",
        "geckodriverpath": "",
        "chromedriverpath": ""
    }

    def setConfig(self, key, value):
        self.appConfig[key] = value

    def getConfig(self):
        return self.appConfig



def getConfig():
    appConfig = AppConfiguration()
    startTime = datetime.now()
    appConfig.setConfig("startTime", startTime)
    print("Process started, logs will be visible in app.log file...")
    logging.info("Start time of execution: " +
                 startTime.strftime('%d_%m_%Y__%H_%M_%S_%f'))
    cwd = os.getcwd()
    print("Select excel file which contains input data for data entry")
    #datafilepath = fileopenbox("Select The Excel Source File: ", "Which contains the input data for data entry",
    #                           None, False)
    datafilepath = fileopenbox("Select The Excel Source File: ", "Which contains the input data for data entry",
                              cwd, False)
    logging.info("Selected file for input data entry: " + datafilepath)
    appConfig.setConfig("datafilepath", datafilepath)
    if platform.system()=='Windows':
        defaultloginfilepath = cwd + "\\login_file.xlsx"
    elif platform.system()=='Darwin':
        defaultloginfilepath = cwd + "/login_file.xlsx"

    #defaultloginfilepath = cwd + "\\login_file.xlsx"
    print(defaultloginfilepath)
    if not path.isfile(defaultloginfilepath):
        print("Select excel(xlsx) file which contains login credential data for data entry")
        loginfilepath = fileopenbox("Select The Excel login credentials File: ",
                                    "Which Contains login credential data for data entry",
                                    cwd, False)
        logging.info("Selected file for login credentials: " + loginfilepath)
    else:
        loginfilepath = defaultloginfilepath

    appConfig.setConfig("loginfilepath", loginfilepath)
    # Open logins file and fetch values
    logins = openpyxl.load_workbook(loginfilepath)
    logins_sheet = logins["login_cred"]

    if logins_sheet.cell(row=2, column=2).value and logins_sheet.cell(row=2, column=2).value:
        u_name = logins_sheet.cell(row=2, column=2).value
        pwd = logins_sheet.cell(row=3, column=2).value
        appConfig.setConfig("u_name", u_name)
        appConfig.setConfig("pwd", pwd)
    else:
        askforlogincred()

    if logins_sheet.cell(row=4, column=2).value:
        prefix = logins_sheet.cell(row=4, column=2).value
        appConfig.setConfig("prefix", prefix)
    else:
        print("No sample ID prefix available for concatenating the Sample ID")
        logging.info(
            "No sample ID prefix available for concatenating the Sample ID")

    if logins_sheet.cell(row=5, column=2).value:
        ProceedWithRATfollowup = logins_sheet.cell(row=5, column=2).value
        appConfig.setConfig("ProceedWithRATfollowup", ProceedWithRATfollowup)

    else:
        print("ProceedWithRATfollowup doesn't have input in the logins sheet")
        logging.info(
            "ProceedWithRATfollowup doesn't have input in the logins sheet")
        appConfig.setConfig("ProceedWithRATfollowup", None)

    if logins_sheet.cell(row=6, column=2).value:
        labUniqueAlert = logins_sheet.cell(row=6, column=2).value
        appConfig.setConfig("labUniqueAlert", labUniqueAlert)
    else:
        labUniqueAlert = None
        print("labUniqueAlert doesn't have input in the logins sheet")
        logging.info("labUniqueAlert doesn't have input in the logins sheet")
        appConfig.setConfig("labUniqueAlert", None)

    if platform.system() !='Darwin':
        if platform.system() != "Linux":
            if platform.machine()=='AMD64':
                geckodriver = cwd + "\\src\\drivers\\win\\64bit\\geckodriver.exe"
                chromedriver = cwd + "\\src\\drivers\\win\\64bit\\chromedriver.exe"

            else:
                geckodriver = cwd + "\\src\\drivers\\win\\32bit\\geckodriver.exe"
                chromedriver = cwd + "\\src\\drivers\\win\\32bit\\chromedriver.exe"
    elif platform.system() =='Darwin' and platform.machine()=='x86_64':
        geckodriver = cwd + "/src/drivers/macos/geckodriver"
        chromedriver = cwd + "/src/drivers/macos/chromedriver"

    while not path.isfile(geckodriver):
        if not logins_sheet.cell(row=7, column=2).value:
            geckodriver = fileopenbox("Select The gecko driver File: ",
                                      "select the Gecko Driver file for Firefox",
                                      cwd, False)
        else:
            geckodriver = logins_sheet.cell(row=7, column=2).value

    while not path.isfile(chromedriver):
        if not logins_sheet.cell(row=7, column=2).value:
            chromedriver = fileopenbox("Select The chrome driver File: ",
                                       "select the chrome Driver file for Google Chrome",
                                       cwd, False)
        else:
            chromedriver = logins_sheet.cell(row=8, column=2).value

    geckodriverString = "Geckodriver path is " + geckodriver
    chromedriverString = "Chromedriver path is " + chromedriver

    print(geckodriverString)
    logging.info(geckodriverString)
    appConfig.setConfig("geckodriverpath", geckodriver)
    print(chromedriverString)
    logging.info(chromedriverString)
    appConfig.setConfig("chromedriverpath", chromedriver)


    if not logins_sheet.cell(row=9, column=2).value:
        driverTy = buttonbox('Select the default Browser' + '\n',
                           "Please Select the default Browser for Chrome or Firefox",
                           ('chrome', 'firefox'))
    elif ((logins_sheet.cell(row=9, column=2).value !="chrome") and (logins_sheet.cell(row=9, column=2).value != "firefox")):
        driverTy = buttonbox('Select the default Browser' + '\n',
                               "Please Select the default Browser for Chrome or Firefox",
                               ('chrome', 'firefox'))
    else:
            driverTy = logins_sheet.cell(row=9, column=2).value

    if driverTy == "chrome":
        print("Selected Chrome as default web driver")
        logging.info("Selected Chrome as default web driver")
        appConfig.setConfig("driverTy", "chrome")
    if driverTy == "firefox":
        print("Selected Firefox as default web driver")
        logging.info("Selected Firefox as default web driver")
        appConfig.setConfig("driverTy", "firefox")
    else:
        driverTy = buttonbox('Select the default Browser' + '\n',
                       "Please Select the default Browser for Chrome or Firefox",
                       ('chrome', 'firefox'))
            #, None, '[<F1>]Yes', '[<F2>]No')
    # initiate browser and open the url
    '''
    driver = "No"
    driver = ynbox('Do you want to change the default Browser to Chrome?' + '\n',
                   "Please click Yes if you wish to change Browser to Chrome",
                   ('[<F1>]Yes', '[<F2>]No'), None, '[<F1>]Yes', '[<F2>]No')
    if driver == "Yes":
        print("Selected Chrome as default web driver")
        logging.info("Selected Chrome as default web driver")
        appConfig.setConfig("driver", "chrome")
    else:
        appConfig.setConfig("driver", "firefox")
    '''
    '''
    if ynbox('Do you want to change the URL of ICMR COVID Portal from: ' + '\n' + appConfig.getConfig()["URL"],
             "Please click Yes if you wish to change or enter the ICMR COVID PORTAL URL",
             ('[<F1>]Yes', '[<F2>]No'), None, '[<F2>]No', '[<F1>]Yes'):
        URL = textbox('Check & Confirm ICMR COVID Portal Web Address',
                      'Please Check and Confirm the ICMR COVID Portal Website Address', appConfig.getConfig()["URL"], 0)
        URL = URL.strip()
        print("Updated ICMR URL to ", URL)
        logging.info("Updated ICMR URL to " + URL)
        appConfig.setConfig("URL", URL)
    '''

